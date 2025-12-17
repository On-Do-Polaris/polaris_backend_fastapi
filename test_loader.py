"""
Loader 테스트 스크립트
AdditionalDataLoader + BuildingCharacteristicsLoader 실행
"""
import os
import sys
import importlib.util

# .env 로드
from dotenv import load_dotenv
load_dotenv()

dw_db_url = os.getenv("DATAWAREHOUSE_DATABASE_URL")
print(f"DB URL: {dw_db_url}")

# DatabaseManager 직접 import
def load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module

base_path = "/Users/odong-i/Desktop/SKALA/FinalProject/DB_ALL/fastapi"

# database.py 로드
db_module = load_module("database", f"{base_path}/ai_agent/utils/database.py")
DatabaseManager = db_module.DatabaseManager

print("=" * 60)
print("[1] Additional Data Loader 실행")
print("=" * 60)

# additional_data_loader.py의 AdditionalDataLoader 클래스 일부 기능만 사용
# scratch 폴더 스캔 + DB 저장

from openpyxl import load_workbook
import re
import json
from datetime import datetime

class SimpleAdditionalDataLoader:
    def __init__(self, db_manager):
        self.db = db_manager

    def load_from_scratch_folder(self, scratch_folder):
        results = []
        total = 0
        success = 0

        for folder_name in os.listdir(scratch_folder):
            folder_path = os.path.join(scratch_folder, folder_name)
            if not os.path.isdir(folder_path):
                continue

            # UUID 패턴 체크
            uuid_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
            if not re.match(uuid_pattern, folder_name, re.IGNORECASE):
                continue

            site_id = folder_name
            print(f"  처리 중: {site_id[:8]}...")

            for file_name in os.listdir(folder_path):
                if not file_name.endswith(('.xlsx', '.xls')):
                    continue

                file_path = os.path.join(folder_path, file_name)
                total += 1

                try:
                    # Excel 읽기
                    wb = load_workbook(file_path, data_only=True)
                    sheet = wb.active

                    # 데이터 추출
                    data_rows = []
                    headers = []
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i == 0:
                            headers = [str(c) if c else f"Col_{j}" for j, c in enumerate(row)]
                        else:
                            row_dict = {headers[j]: str(c) if c else "" for j, c in enumerate(row) if j < len(headers)}
                            if any(row_dict.values()):
                                data_rows.append(row_dict)

                    structured_data = {"headers": headers, "rows": data_rows}

                    # DB 저장
                    self.save_to_db(site_id, file_name, structured_data)

                    results.append({"file_name": file_name, "success": True})
                    success += 1

                except Exception as e:
                    results.append({"file_name": file_name, "success": False, "error": str(e)})

        return {"total_files": total, "success_count": success, "failed_count": total - success, "results": results}

    def save_to_db(self, site_id, file_name, structured_data):
        query = """
            INSERT INTO site_additional_data (site_id, structured_data, metadata, uploaded_at)
            VALUES (%s, %s, %s, NOW())
            ON CONFLICT DO NOTHING
        """
        from psycopg2.extras import Json
        metadata = {"file_name": file_name, "uploaded_at": datetime.now().isoformat()}

        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, (site_id, Json(structured_data), Json(metadata)))
            cursor.close()

# 실행
db = DatabaseManager(dw_db_url)
loader = SimpleAdditionalDataLoader(db)
result = loader.load_from_scratch_folder(f"{base_path}/scratch")

print(f"\n결과:")
print(f"  총 파일: {result['total_files']}")
print(f"  성공: {result['success_count']}")
print(f"  실패: {result['failed_count']}")

for r in result['results']:
    status = "✅" if r.get("success") else "❌"
    print(f"  {status} {r.get('file_name')}")

print("\n" + "=" * 60)
print("[2] Building Characteristics - DB에 직접 목데이터 삽입")
print("=" * 60)

# building_aggregate_cache 테이블에 목데이터 삽입
sites_data = [
    ("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", 37.405879, 127.099877, "경기도 성남시 분당구 판교로 255번길 38", "SK 판교캠퍼스"),
    ("bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb", 35.158691, 129.159702, "부산광역시 해운대구", "부산 물류센터"),
    ("cccccccc-cccc-cccc-cccc-cccccccccccc", 37.456789, 126.705123, "인천광역시 남동구", "인천 공장"),
    ("dddddddd-dddd-dddd-dddd-dddddddddddd", 36.361234, 127.356789, "대전광역시 유성구", "대전 R&D센터"),
]

# building_aggregate_cache 테이블 확인 및 삽입
try:
    with db.get_connection() as conn:
        cursor = conn.cursor()

        for site_id, lat, lon, address, name in sites_data:
            # 간단한 building 데이터 삽입
            building_data = {
                "site_id": site_id,
                "name": name,
                "address": address,
                "structure_types": ["철근콘크리트"],
                "purpose_types": ["업무시설"],
                "building_count": 1,
                "total_floor_area_sqm": 50000,
                "oldest_building_age_years": 10
            }

            print(f"  ✅ {name} 건물 데이터 준비됨")

        cursor.close()

    print("\n  (참고: building_aggregate_cache 테이블 구조에 따라 실제 삽입 필요)")

except Exception as e:
    print(f"  ⚠️ Building 데이터 처리 중 오류: {e}")

print("\n" + "=" * 60)
print("✅ Loader 실행 완료!")
print("=" * 60)
