'''
파일명: additional_data_loader.py
작성일: 2025-12-16
버전: v05 (Key-Value 파싱 + data_category 제거)
파일 개요: 추가 데이터 (Excel) ETL 로더

변경 내역 (v05):
    - Excel 파싱 방식 개선: 텍스트 덤프 → Key-Value 구조
    - 파싱 실패 시 자동 폴백 (Fallback 모드)
    - data_category 필드 제거 (metadata로 이동)

================================================================================
[ 입력 데이터 요구사항 ]
================================================================================

1. load_from_scratch_folder() - scratch 폴더에서 Excel 파일 → DB 적재
   입력:
       - scratch_folder: str (scratch 폴더 경로)
         예: "/path/to/scratch"

   폴더 구조:
       scratch/
       ├── 22222222-2222-2222-2222-222222222222/   ← 폴더명 = site_id
       │   ├── 에너지_사용량.xlsx
       │   └── 전력_사용량.xlsx
       └── 44444444-4444-4444-4444-444444444444/
           └── ...

   출력:
       - Dict: 적재 결과
         {
           "total_files": 5,
           "success_count": 5,
           "failed_count": 0,
           "results": [...]
         }

2. load_and_save() - 단일 Excel 파일 → DB 적재
   입력:
       - file_path: str (Excel 파일 경로, 필수)
       - site_id: str (사업장 UUID, 필수)
       - category: str (선택, metadata에 저장됨)

3. fetch_all_for_site() - DB에서 사이트별 추가 데이터 조회
   입력:
       - site_id: str (사업장 UUID, 필수)

   출력:
       - Dict[str, List]: 카테고리별 데이터

================================================================================
[ DB 테이블 ]
================================================================================

site_additional_data (datawarehouse DB):
    - PK: id (UUID)
    - 주요 컬럼:
        - site_id: UUID (FK → sites.id)
        - raw_text: text (PDF 추출용)
        - structured_data: JSONB (Key-Value 파싱된 Excel 데이터)
        - file_content: JSONB (파일 내용)
        - metadata: JSONB (파일 메타 정보, 카테고리 포함)
        - uploaded_by: uuid
        - uploaded_at: timestamp
        - expires_at: timestamp

    주의: data_category 컬럼 제거됨 (metadata.inferred_category로 대체)

================================================================================
[ 파싱 방식 ]
================================================================================

1. Structured 모드 (기본):
    - 첫 행을 헤더로 인식
    - 각 행을 딕셔너리로 변환
    - JSON 크기 최적화

2. Fallback 모드 (파싱 실패 시):
    - 텍스트 덤프 방식 사용
    - 모든 셀 값을 " | "로 연결

3. 예외 처리:
    - 빈 헤더 → "Column_N" 자동 생성
    - 중복 헤더 → "_2", "_3" 접미사 추가
    - 빈 행 자동 스킵

================================================================================
[ 카테고리 자동 추론 규칙 ]
================================================================================

파일명에 포함된 키워드로 카테고리 추론 (metadata.inferred_category):
    - "전력" 또는 "power" → "power"
    - "에너지" 또는 "energy" → "energy"
    - "보험" 또는 "insurance" → "insurance"
    - "건물" 또는 "building" → "building"
    - "자산" 또는 "asset" → "asset"
    - "시설" 또는 "facility" → "facility"
    - 그 외 → "other"

================================================================================
'''

from typing import Dict, Any, List, Optional
import logging
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# DatabaseManager 임포트
try:
    from ...utils.database import DatabaseManager
except ImportError:
    DatabaseManager = None
    print("⚠️ DatabaseManager를 임포트할 수 없습니다.")

logger = logging.getLogger(__name__)


class AdditionalDataLoader:
    """
    추가 데이터 ETL 로더 (Excel → DB 적재)

    역할:
        1. scratch 폴더 스캔 (폴더명 = site_id)
        2. Excel 파일 읽기 (범용 텍스트 덤프)
        3. site_additional_data 테이블에 적재

    사용 예시:
        loader = AdditionalDataLoader(db_url="postgresql://...")

        # 방법 1: scratch 폴더 전체 적재
        loader.load_from_scratch_folder("/path/to/scratch")

        # 방법 2: 단일 파일 적재
        loader.load_and_save("path/to/file.xlsx", site_id="uuid-...")
    """

    def __init__(self, db_url: Optional[str] = None, db_manager=None):
        """
        초기화

        Args:
            db_url: Datawarehouse DB URL (site_additional_data 테이블 접근용) - DEPRECATED
            db_manager: DatabaseManager 인스턴스 (권장)
        """
        self.logger = logger

        # DatabaseManager 초기화 또는 전달받은 인스턴스 사용
        if db_manager:
            # 전달받은 DatabaseManager 인스턴스 사용 (권장)
            self.db_manager = db_manager
            self.logger.info("DatabaseManager 인스턴스 전달받음 (site_additional_data)")
        elif DatabaseManager and db_url:
            # DEPRECATED: URL로 초기화 (하위 호환성)
            self.logger.warning("db_url 사용은 deprecated입니다. db_manager를 전달하세요.")
            self.db_manager = None
        else:
            self.db_manager = None
            self.logger.warning("DatabaseManager 없음 - DB 적재 비활성화")

        self.logger.info("AdditionalDataLoader 초기화 완료")

    # ==========================================================================
    # scratch 폴더 기반 적재 (폴더명 = site_id)
    # ==========================================================================

    def load_from_scratch_folder(self, scratch_folder: str) -> Dict[str, Any]:
        """
        scratch 폴더 내 모든 Excel 파일을 DB에 적재

        폴더 구조:
            scratch/
            ├── {site_id_1}/
            │   ├── file1.xlsx
            │   └── file2.xlsx
            └── {site_id_2}/
                └── file3.xlsx

        Args:
            scratch_folder: scratch 폴더 경로

        Returns:
            적재 결과 요약
        """
        self.logger.info(f"scratch 폴더 스캔 시작: {scratch_folder}")

        if not os.path.exists(scratch_folder):
            self.logger.error(f"scratch 폴더가 존재하지 않음: {scratch_folder}")
            return {"error": "폴더 없음", "total_files": 0}

        results = []
        total_files = 0
        success_count = 0
        failed_count = 0

        # 하위 폴더 순회 (폴더명 = site_id)
        for folder_name in os.listdir(scratch_folder):
            folder_path = os.path.join(scratch_folder, folder_name)

            if not os.path.isdir(folder_path):
                continue

            # 폴더명이 UUID 형식인지 확인
            site_id = self._extract_site_id_from_folder(folder_name)
            if not site_id:
                self.logger.warning(f"UUID 형식이 아닌 폴더 스킵: {folder_name}")
                continue

            self.logger.info(f"사업장 폴더 처리: {folder_name} → site_id={site_id[:8]}...")

            # 폴더 내 Excel 파일 처리
            for file_name in os.listdir(folder_path):
                if not file_name.endswith(('.xlsx', '.xls')):
                    continue

                file_path = os.path.join(folder_path, file_name)
                total_files += 1

                result = self.load_and_save(file_path, site_id)
                results.append(result)

                if result.get("success"):
                    success_count += 1
                else:
                    failed_count += 1

        summary = {
            "total_files": total_files,
            "success_count": success_count,
            "failed_count": failed_count,
            "deleted_files": sum(1 for r in results if r.get("file_deleted")),
            "results": results
        }

        self.logger.info(
            f"scratch 폴더 적재 완료: {success_count}/{total_files} 성공, "
            f"{summary['deleted_files']}개 파일 삭제"
        )
        return summary

    def _extract_site_id_from_folder(self, folder_name: str) -> Optional[str]:
        """
        폴더명에서 site_id (UUID) 추출

        Args:
            folder_name: 폴더명

        Returns:
            UUID 문자열 또는 None
        """
        # UUID 패턴 (8-4-4-4-12)
        uuid_pattern = r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$'

        if re.match(uuid_pattern, folder_name):
            return folder_name.lower()  # 소문자로 정규화

        return None

    # ==========================================================================
    # 단일 파일 적재
    # ==========================================================================

    def load_and_save(
        self,
        file_path: str,
        site_id: str,
        category: str = None,  # 호환성 유지 (metadata로 저장)
        auto_cleanup: bool = True  # 자동 삭제 플래그 (성공 시 파일 삭제)
    ) -> Dict[str, Any]:
        """
        Excel 파일을 읽어서 site_additional_data 테이블에 적재

        Args:
            file_path: Excel 파일 경로
            site_id: 사업장 UUID
            category: 데이터 카테고리 (선택, metadata에 저장됨)

        Returns:
            적재 결과 정보
        """
        file_name = Path(file_path).name
        self.logger.info(f"Excel → DB 적재: {file_name}")

        if not self.db_manager:
            self.logger.error("DB 연결 없음 - 적재 불가")
            return {"success": False, "error": "DB 연결 없음", "file_name": file_name}

        try:
            # 1. Excel 파일을 Key-Value 구조로 파싱
            data = self._universal_excel_dump(file_path)

            if 'error' in data:
                self.logger.error(f"Excel 파일 읽기 실패: {data['error']}")
                return {"success": False, "error": data['error'], "file_name": file_name}

            # 2. 카테고리 추론 (metadata용)
            if not category:
                category = self._infer_category(file_name)

            # 3. 메타데이터 구성
            metadata = {
                'source': 'AdditionalDataLoader',
                'file_name': file_name,
                'inferred_category': category,  # 카테고리는 metadata에 저장
                'parsing_method': data.get('parsing_method', 'unknown'),
                'loaded_at': datetime.now().isoformat(),
                'sheet_count': len(data.get('sheets', [])),
                'total_rows': sum(s.get('row_count', 0) for s in data.get('sheets', []))
            }

            # 4. DB에 저장
            self.db_manager.save_additional_data(
                site_id=site_id,
                structured_data=data,
                metadata=metadata
            )

            self.logger.info(f"✅ DB 적재 완료: {file_name} → site_id={site_id[:8]}..., parsing={data.get('parsing_method')}")

            # 5. Auto cleanup (성공 시 파일 삭제)
            cleanup_status = None
            if auto_cleanup:
                cleanup_status = self.cleanup_file(file_path)
                if cleanup_status:
                    self.logger.info(f"🗑️ 파일 자동 삭제 완료: {file_path}")
                else:
                    self.logger.warning(f"⚠️ 파일 삭제 실패: {file_path}")

            return {
                "success": True,
                "site_id": site_id,
                "category": category,  # 호환성 유지
                "parsing_method": data.get('parsing_method'),
                "file_name": file_name,
                "sheet_count": len(data.get('sheets', [])),
                "total_rows": metadata['total_rows'],
                "file_deleted": cleanup_status  # 삭제 여부
            }

        except Exception as e:
            self.logger.error(f"DB 적재 실패: {e}")
            return {"success": False, "error": str(e), "file_name": file_name}

    def _universal_excel_dump(self, file_path: str) -> Dict[str, Any]:
        """
        Excel을 Key-Value 구조로 파싱 (폴백: 텍스트 덤프)

        파싱 전략:
            1. 첫 행을 헤더로 인식
            2. 각 행을 딕셔너리로 변환
            3. 파싱 실패 시 텍스트 덤프로 폴백

        Args:
            file_path: Excel 파일 경로

        Returns:
            파싱된 데이터 딕셔너리 (structured 또는 fallback 모드)
        """
        result = {
            'file_name': Path(file_path).name,
            'uploaded_at': datetime.now().isoformat(),
            'sheets': [],
            'parsing_method': 'unknown'
        }

        try:
            wb = load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 시트 파싱 시도
                sheet_data = self._parse_sheet_structured(ws, sheet_name)

                # 파싱 실패 시 폴백 모드
                if sheet_data.get('parsing_failed'):
                    sheet_data = self._parse_sheet_fallback(ws, sheet_name)
                    result['parsing_method'] = 'fallback'
                else:
                    result['parsing_method'] = 'structured'

                result['sheets'].append(sheet_data)

            wb.close()

        except Exception as e:
            self.logger.error(f"Excel 파일 읽기 실패: {e}")
            result['error'] = str(e)
            result['sheets'] = []
            result['parsing_method'] = 'error'

        return result

    def _parse_sheet_structured(self, ws, sheet_name: str) -> Dict[str, Any]:
        """
        시트를 Key-Value 구조로 파싱

        Args:
            ws: openpyxl Worksheet 객체
            sheet_name: 시트명

        Returns:
            파싱된 시트 데이터 또는 실패 플래그
        """
        try:
            all_rows = list(ws.iter_rows(values_only=True))

            if not all_rows or len(all_rows) < 2:
                # 데이터가 없거나 헤더만 있으면 실패
                return {'parsing_failed': True}

            # Step 1: 첫 행을 헤더로 간주
            header_row = all_rows[0]
            headers = []
            for idx, cell_value in enumerate(header_row):
                if cell_value is None or str(cell_value).strip() == '':
                    headers.append(f'Column_{idx + 1}')  # 빈 헤더는 자동 이름
                else:
                    headers.append(str(cell_value).strip())

            # 중복 헤더 처리 (같은 이름에 _2, _3 추가)
            seen = {}
            for i, header in enumerate(headers):
                if header in seen:
                    seen[header] += 1
                    headers[i] = f"{header}_{seen[header]}"
                else:
                    seen[header] = 1

            # Step 2: 데이터 행을 딕셔너리로 변환
            data_rows = []
            for row_values in all_rows[1:]:
                # 완전히 빈 행은 스킵
                if not any(v is not None and str(v).strip() != '' for v in row_values):
                    continue

                row_dict = {}
                for idx, value in enumerate(row_values):
                    if idx < len(headers):
                        # 값 타입 자동 변환 (숫자는 숫자로 유지)
                        if value is None:
                            row_dict[headers[idx]] = None
                        elif isinstance(value, (int, float)):
                            row_dict[headers[idx]] = value
                        else:
                            row_dict[headers[idx]] = str(value).strip()

                data_rows.append(row_dict)

            # Step 3: 결과 반환
            return {
                'name': sheet_name,
                'headers': headers,
                'data': data_rows,
                'row_count': len(data_rows),
                'parsing_method': 'structured'
            }

        except Exception as e:
            self.logger.warning(f"시트 '{sheet_name}' 파싱 실패: {e}, 폴백 모드로 전환")
            return {'parsing_failed': True}

    def _parse_sheet_fallback(self, ws, sheet_name: str) -> Dict[str, Any]:
        """
        폴백 모드: 텍스트 덤프 방식

        Args:
            ws: openpyxl Worksheet 객체
            sheet_name: 시트명

        Returns:
            텍스트 덤프된 시트 데이터
        """
        rows = []
        for row in ws.iter_rows():
            row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
            # 완전히 빈 행은 스킵
            if any(v.strip() for v in row_values):
                rows.append(' | '.join(row_values))

        return {
            'name': sheet_name,
            'row_count': len(rows),
            'content': '\n'.join(rows),
            'parsing_method': 'fallback'
        }

    def _infer_category(self, file_name: str) -> str:
        """파일명에서 카테고리 추론"""
        file_lower = file_name.lower()

        if '전력' in file_lower or 'power' in file_lower:
            return 'power'
        elif '에너지' in file_lower or 'energy' in file_lower:
            return 'energy'
        elif '보험' in file_lower or 'insurance' in file_lower:
            return 'insurance'
        elif '건물' in file_lower or 'building' in file_lower:
            return 'building'
        elif '자산' in file_lower or 'asset' in file_lower:
            return 'asset'
        elif '시설' in file_lower or 'facility' in file_lower:
            return 'facility'
        else:
            return 'other'

    # ==========================================================================
    # DB 조회 메서드 (Agent가 사용)
    # ==========================================================================

    def fetch_from_db(
        self,
        site_id: str,
        category: str = None  # 호환성 유지 (metadata 필터링)
    ) -> List[Dict[str, Any]]:
        """
        site_additional_data 테이블에서 데이터 조회

        Args:
            site_id: 사업장 UUID
            category: 데이터 카테고리 (선택, metadata에서 필터링)

        Returns:
            조회된 데이터 리스트
        """
        if not self.db_manager:
            self.logger.warning("DB 연결 없음 - 조회 불가")
            return []

        try:
            # data_category 필드 제거됨, site_id만으로 조회
            results = self.db_manager.fetch_additional_data(site_id=site_id)

            # 카테고리 필터링 (metadata에서)
            if category:
                filtered_results = []
                for item in results:
                    metadata = item.get('metadata', {})
                    if metadata.get('inferred_category') == category:
                        filtered_results.append(item)
                results = filtered_results

            self.logger.info(f"DB 조회 완료: site_id={site_id[:8]}..., {len(results)}개 데이터")
            return results

        except Exception as e:
            self.logger.error(f"DB 조회 실패: {e}")
            return []

    def fetch_all_for_site(self, site_id: str) -> Dict[str, Any]:
        """
        특정 사업장의 모든 추가 데이터를 카테고리별로 정리하여 반환

        Args:
            site_id: 사업장 UUID

        Returns:
            카테고리별 데이터 딕셔너리
            {
                "energy": [...],
                "power": [...],
                ...
            }
        """
        all_data = self.fetch_from_db(site_id)

        # metadata에서 카테고리 추출하여 정리
        categorized = {}
        for item in all_data:
            metadata = item.get('metadata', {})
            cat = metadata.get('inferred_category', 'other')

            if cat not in categorized:
                categorized[cat] = []
            categorized[cat].append(item)

        return categorized

    # ==========================================================================
    # 유틸리티
    # ==========================================================================

    def cleanup_file(self, file_path: str) -> bool:
        """
        Excel 파일 삭제 (분석 완료 후 정리)

        Args:
            file_path: 삭제할 Excel 파일 경로

        Returns:
            삭제 성공 여부
        """
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                self.logger.info(f"✅ Excel 파일 삭제 완료: {file_path}")
                return True
            else:
                self.logger.warning(f"⚠️ Excel 파일이 존재하지 않음: {file_path}")
                return False
        except Exception as e:
            self.logger.error(f"❌ Excel 파일 삭제 실패: {file_path}, 오류: {e}")
            return False
